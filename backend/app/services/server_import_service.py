from __future__ import annotations

import hashlib
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.exceptions import AppError
from app.models.asset import ModelAsset, ModelVersion
from app.models.borehole import Borehole, BoreholeSegment
from app.models.project import Project
from app.models.workspace import ImportRun, WorkingFace


LAYER_COLORS = (
    "#8d7358",
    "#c7a66b",
    "#65755f",
    "#9b7d68",
    "#6f7f89",
    "#b68b62",
    "#74665a",
)


def normalize_match_key(value: Any) -> str:
    return "".join(char for char in str(value or "").strip().upper() if char.isalnum())


def canonical_code(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper()


def as_float(value: Any, default: float = 0) -> float:
    if value in (None, ""):
        return default
    return float(value)


def rows_as_dicts(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = [str(value).strip() for value in next(rows)]
        return [dict(zip(headers, row, strict=False)) for row in rows if any(row)]
    finally:
        workbook.close()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


class ServerDataImportService:
    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.data_root = settings.source_data_path
        self.model_root = settings.source_model_path

    async def run(self, project_name: str) -> dict[str, Any]:
        self._validate_sources()
        locations = self._read_locations()
        strata = self._read_strata()
        project = await self._upsert_project(project_name, locations)
        borehole_summary = await self._upsert_boreholes(project, locations, strata)
        model_count = await self._upsert_models(project)
        working_face_count = await self._upsert_working_faces(project)

        summary = {
            "project_id": project.id,
            **borehole_summary,
            "model_assets": model_count,
            "working_faces": working_face_count,
        }
        self.session.add(
            ImportRun(
                project_id=project.id,
                source="server/data + server/static/models",
                status="completed",
                summary_json=summary,
            )
        )
        await self.session.commit()
        return summary

    def _validate_sources(self) -> None:
        required = (
            self.data_root / "location" / "钻孔位置.xlsx",
            self.data_root / "boreholes" / "地层汇总.xlsx",
            self.data_root / "workingfaces.json",
            self.model_root,
        )
        missing = [str(path) for path in required if not path.exists()]
        if missing:
            raise AppError("IMPORT_SOURCE_MISSING", f"数据源不存在: {', '.join(missing)}")

    def _read_locations(self) -> dict[str, dict[str, Any]]:
        path = self.data_root / "location" / "钻孔位置.xlsx"
        result: dict[str, dict[str, Any]] = {}
        for row in rows_as_dicts(path):
            key = normalize_match_key(row.get("name"))
            if not key:
                continue
            result[key] = {
                "code": canonical_code(row["name"]),
                "x": as_float(row.get("x")),
                "y": as_float(row.get("y")),
                "z": as_float(row.get("z")),
            }
        return result

    def _read_strata(self) -> dict[str, list[dict[str, Any]]]:
        path = self.data_root / "boreholes" / "地层汇总.xlsx"
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows_as_dicts(path):
            key = normalize_match_key(row.get("钻孔名称"))
            if not key:
                continue
            bottom = as_float(row.get("深度"))
            thickness = as_float(row.get("厚度"))
            grouped[key].append(
                {
                    "source_code": canonical_code(row.get("钻孔名称")),
                    "layer_name": str(row.get("地层名称") or "未命名地层").strip(),
                    "top_depth": bottom - thickness,
                    "bottom_depth": bottom,
                    "thickness": thickness,
                }
            )
        return grouped

    async def _upsert_project(
        self, project_name: str, locations: dict[str, dict[str, Any]]
    ) -> Project:
        project = await self.session.scalar(select(Project).where(Project.name == project_name))
        coordinates = list(locations.values())
        origin_x = sum(item["x"] for item in coordinates) / len(coordinates)
        origin_y = sum(item["y"] for item in coordinates) / len(coordinates)
        origin_z = sum(item["z"] for item in coordinates) / len(coordinates)
        if project is None:
            project = Project(name=project_name)
            self.session.add(project)
        project.description = "由 server 目录地质模型与钻孔资料构建的三维地质工作区"
        project.coordinate_system = "Gauss-Kruger"
        project.origin_x = origin_x
        project.origin_y = origin_y
        project.origin_z = origin_z
        # 现有静态模型生成流程已对高程使用 20 倍夸张，钻孔必须采用同一比例。
        project.vertical_scale = 20
        await self.session.flush()
        return project

    async def _upsert_boreholes(
        self,
        project: Project,
        locations: dict[str, dict[str, Any]],
        strata: dict[str, list[dict[str, Any]]],
    ) -> dict[str, int]:
        existing_rows = await self.session.scalars(
            select(Borehole).where(Borehole.project_id == project.id)
        )
        existing = {normalize_match_key(item.code): item for item in existing_rows}
        segment_count = 0
        zero_thickness = 0

        for key in sorted(set(locations) | set(strata)):
            location = locations.get(key)
            segments = strata.get(key, [])
            code = location["code"] if location else segments[0]["source_code"]
            borehole = existing.get(key)
            if borehole is None:
                borehole = Borehole(project_id=project.id, code=code, name=code, x=0, y=0, z=0)
                self.session.add(borehole)
                await self.session.flush()
            borehole.code = code
            borehole.name = code
            borehole.x = location["x"] if location else 0
            borehole.y = location["y"] if location else 0
            borehole.z = location["z"] if location else 0
            borehole.total_depth = max((item["bottom_depth"] for item in segments), default=0)
            borehole.status = "active" if location else "missing_location"
            borehole.metadata_json = {
                "source": "server/data",
                "has_location": location is not None,
                "has_strata": bool(segments),
            }

            await self.session.execute(
                delete(BoreholeSegment).where(BoreholeSegment.borehole_id == borehole.id)
            )
            for sequence, item in enumerate(segments):
                zero_thickness += int(item["thickness"] == 0)
                self.session.add(
                    BoreholeSegment(
                        borehole_id=borehole.id,
                        layer_name=item["layer_name"],
                        lithology=item["layer_name"],
                        top_depth=item["top_depth"],
                        bottom_depth=item["bottom_depth"],
                        thickness=item["thickness"],
                        color=LAYER_COLORS[sequence % len(LAYER_COLORS)],
                        sequence=sequence,
                    )
                )
            segment_count += len(segments)

        return {
            "boreholes": len(set(locations) | set(strata)),
            "borehole_segments": segment_count,
            "zero_thickness_segments": zero_thickness,
        }

    async def _upsert_models(self, project: Project) -> int:
        metadata_path = self.data_root / "models_meta.json"
        metadata_rows = json.loads(metadata_path.read_text(encoding="utf-8")) if metadata_path.exists() else []
        metadata_by_file = {item.get("fileName"): item for item in metadata_rows}
        existing_rows = await self.session.scalars(
            select(ModelAsset).where(ModelAsset.project_id == project.id)
        )
        existing = {
            item.metadata_json.get("source_file"): item
            for item in existing_rows
            if item.metadata_json.get("source_file")
        }
        model_paths = sorted(self.model_root.rglob("*.glb"))

        for path in model_paths:
            relative_path = path.relative_to(self.model_root).as_posix()
            metadata = metadata_by_file.get(path.name, {})
            asset = existing.get(relative_path)
            if asset is None:
                asset = ModelAsset(project_id=project.id, name=metadata.get("name") or path.stem)
                self.session.add(asset)
            asset.name = metadata.get("name") or path.stem
            asset.model_type = metadata.get("type") or "stratum"
            asset.status = "ready"
            asset.metadata_json = {
                "source": "server/static/models",
                "source_file": relative_path,
                "description": metadata.get("description"),
                "format": "glb",
                "bbox": metadata.get("bbox"),
            }
            await self.session.flush()

            version = await self.session.scalar(
                select(ModelVersion).where(
                    ModelVersion.model_id == asset.id,
                    ModelVersion.version == 1,
                )
            )
            if version is None:
                version = ModelVersion(model_id=asset.id, version=1)
                self.session.add(version)
            version.file_path = relative_path
            version.storage_scope = "server_static"
            version.file_size = path.stat().st_size
            version.content_hash = file_sha256(path)
            version.draco_compressed = False
            await self.session.flush()
            asset.current_version_id = version.id

        return len(model_paths)

    async def _upsert_working_faces(self, project: Project) -> int:
        rows = json.loads((self.data_root / "workingfaces.json").read_text(encoding="utf-8"))
        existing_rows = await self.session.scalars(
            select(WorkingFace).where(WorkingFace.project_id == project.id)
        )
        existing = {item.code: item for item in existing_rows}
        for row in rows:
            code = str(row["code"])
            item = existing.get(code)
            if item is None:
                item = WorkingFace(project_id=project.id, code=code, name=row["name"])
                self.session.add(item)
            item.name = row["name"]
            item.status = row.get("status") or "规划中"
            item.description = row.get("description")
            item.length = row.get("length")
            item.width = row.get("width")
            item.coal_seam = row.get("coalSeam")
            item.metadata_json = {"legacy_id": row.get("id"), "legacy_model_id": row.get("modelId")}
        return len(rows)
